import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from django.shortcuts import render, redirect
from django.http import JsonResponse
from openpyxl import Workbook, load_workbook
from django.conf import settings
from django.contrib import messages


def index(request):
    request.session.flush()
    return render(request, "frontend/index.html")


def submit_form(request):
    if request.method == "POST":
        try:
            post_data = request.POST

            # excel_file_path = 'media/Spirit-Of-Diwali.xlsx'
            excel_file_path = os.path.join(settings.MEDIA_ROOT, 'excel/Spirit-Of-Diwali.xlsx')

            try:
                workbook = load_workbook(excel_file_path)
            except FileNotFoundError:
                workbook = Workbook()

            sheet_name = "Vendor Registrations"

            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(title=sheet_name)

            # Find the header row (assuming it's the first row)
            header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

            # Get the column index based on the existing column names
            column_index = {
                "Name": header_row.index("Name") + 1,
                "Phone No.": header_row.index("Phone No.") + 1,
                "Email": header_row.index("Email") + 1
            }

            # Prepare the data to append
            data_to_append = [
                post_data.get("name", ""),
                post_data.get("number", ""),
                post_data.get("email", "")
            ]

            # Append the data to the corresponding columns in a new row
            new_row = worksheet.max_row + 1
            for col_name, value in zip(column_index.keys(), data_to_append):
                col_index = column_index[col_name]
                worksheet.cell(row=new_row, column=col_index, value=value)

            # Save the changes
            workbook.save(excel_file_path)

            # calling send_email function
            # send_email(data_to_append)
            try:
                subject = "Coming to the Event"
                body = f"Name: {data_to_append[0]}\nPhone No.: {data_to_append[1]}\nEmail: {data_to_append[2]}"

                # Set up the email
                sender_email = settings.EMAIL_HOST_USER
                sender_password = settings.EMAIL_HOST_PASSWORD
                # receiver_emails = ["kartik@tailwebs.com"]
                receiver_emails = [
                    "diwalimelbourne@gmail.com", "spiritevents9@gmail.com",
                    "vedika@tailwebs.com", "kartik@tailwebs.com"
                ]

                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = ", ".join(receiver_emails)
                msg['Subject'] = subject
                msg.attach(MIMEText(body, 'plain'))

                # Connect to the SMTP server and send the email
                server = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
                server.starttls()
                server.login(sender_email, sender_password)
                server.sendmail(sender_email, receiver_emails, msg.as_string())
                server.quit()

                print("===> Email has been successfully delivered...")
            except Exception as e:
                return JsonResponse(
                    {"error for sending email: ": str(e)}, status=500
                )

            request.session["success"] = post_data.get("name", "")
            messages.success(request, 'Profile details added.')
            return redirect('/success')
        except Exception as e:
            return JsonResponse(
                {"error": str(e)}, status=500
            )
    else:
        return JsonResponse(
            {"message": "Only POST requests are allowed"}, status=405
        )


def success(request):
    if "success" not in request.session:
        return redirect('/')
    else:
        return render(request, "frontend/success.html")
